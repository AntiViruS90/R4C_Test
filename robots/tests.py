from django.db import DatabaseError
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from unittest.mock import patch
import json

import io
from openpyxl import Workbook, load_workbook

from customers.models import Customer
from robots.models import Robot
from orders.models import Order
from robots.signals import notify_customers_when_robot_available


def load_workbook_from_response(response):
    """Вспомогательная функция для загрузки книги из ответа"""
    return Workbook(io.BytesIO(response.content))


def check_workbook_headers(ws):
    """Вспомогательная функция для проверки заголовков в листе"""
    assert ws.cell(row=1, column=1).value == 'Модель'
    assert ws.cell(row=1, column=2).value == 'Версия'
    assert ws.cell(row=1, column=3).value == 'Количество за неделю'


def check_data_rows(ws, expected_data):
    """Вспомогательная функция для проверки данных в строках листа"""
    for row, (model, version, count) in enumerate(expected_data, start=2):
        assert ws.cell(row=row, column=1).value == model
        assert ws.cell(row=row, column=2).value == version
        assert ws.cell(row=row, column=3).value == count


class CreateRobotTestCase(TestCase):

    def test_create_robot_success(self):
        url = reverse('create_robot')
        data = {
            "model": "XR",
            "version": "10",
            "created": timezone.now().isoformat()
        }
        response = self.client.post(
            url,
            data=json.dumps(data),
            content_type='application/json'
        )

        self.assertEqual(response.status_code, 201)
        self.assertIn('message', response.json())
        self.assertIn('id', response.json())
        self.assertEqual(
            response.json()['message'],
            'Robot created succesfully'
        )

    def test_create_robot_missing_fields(self):
        url = reverse('create_robot')

        data = {
            "version": "10",
            "created": timezone.now().isoformat()
        }
        response = self.client.post(
            url,
            data=json.dumps(data),
            content_type='application/json'
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'Missing required fields')

    def test_create_robot_invalid_date_format(self):
        url = reverse('create_robot')

        data = {
            "model": "XR",
            "version": "10",
            "created": "invalid_date_format"
        }
        response = self.client.post(
            url,
            data=json.dumps(data),
            content_type='application/json'
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'Invalid date format')

    def test_create_robot_invalid_json(self):
        url = reverse('create_robot')

        data = "{model: XR, version: 10, created: '2024-12-12T00:00:00'}"
        response = self.client.post(url, data=data, content_type='application/json')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'Invalid JSON format')

    def test_create_robot_database_error(self):
        url = reverse('create_robot')

        data = {
            "model": "XR",
            "version": "10",
            "created": timezone.now().isoformat()
        }

        with patch('robots.models.Robot.objects.create', side_effect=DatabaseError("Database is down")):
            response = self.client.post(url, data=json.dumps(data), content_type='application/json')

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.json()['error'], "Database error: Database is down")

    def test_create_robot_invalid_method(self):
        url = reverse('create_robot')

        response = self.client.get(url)

        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.json()['error'], "Only POST method is allowed")

    def test_generate_robot_summary_no_robots(self):
        url = reverse('download_robots_summary')
        with patch('robots.models.Robot.objects.filter') as mock_filter:
            mock_filter.return_value = Robot.objects.none()
            response = self.client.get(url)

        assert response.status_code == 200
        wb = load_workbook_from_response(response)
        assert len(wb.sheetnames) == 0

    def test_generate_robot_summary_single_model(self):
        url = reverse('download_robots_summary')
        Robot.objects.create(model='R2', version='10', created=timezone.now())
        Robot.objects.create(model='R2', version='11', created=timezone.now())
        response = self.client.get(url)

        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))

        assert len(wb.sheetnames) == 2
        assert 'R2' in wb.sheetnames

        ws = wb['R2']
        check_workbook_headers(ws)
        check_data_rows(ws, [('R2', '10', 1), ('R2', '11', 1)])

    def test_generate_robot_summary_multiple_models(self):
        url = reverse('download_robots_summary')
        Robot.objects.create(model='R2', version='10', created=timezone.now())
        Robot.objects.create(model='R3', version='10', created=timezone.now())
        response = self.client.get(url)

        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        assert len(wb.sheetnames) == 3
        assert 'R2' in wb.sheetnames
        assert 'R3' in wb.sheetnames

        ws1, ws2 = wb['R2'], wb['R3']
        check_workbook_headers(ws1)
        check_workbook_headers(ws2)

        check_data_rows(ws1, [('R2', '10', 1)])
        check_data_rows(ws2, [('R3', '10', 1)])

    def test_notify_customers_when_robot_available_no_serial(self):
        robot = Robot.objects.create(model='XR', version='10', created=timezone.now(), serial='XR10')

        with patch('orders.models.Order.objects.filter') as mock_filter:
            mock_filter.return_value = Order.objects.none()

            notify_customers_when_robot_available(Robot, robot, created=True)

            mock_filter.assert_called_once_with(robot_serial='XR10', is_waiting=True)

    def test_notify_customers_when_robot_available_email_failure(self):
        customer = Customer.objects.create(email='customer@example.com')
        robot = Robot.objects.create(model='XR', version='10', created=timezone.now(), serial='XR10')
        order = Order.objects.create(robot_serial='XR10', is_waiting=True, customer=customer)

        with patch('robots.signals.send_email_to_customer',
                   side_effect=Exception("Email sending failed")) as mock_send_email:
            notify_customers_when_robot_available(Robot, robot, created=True)

            mock_send_email.assert_called_once_with(order.customer, robot)
            order.refresh_from_db()
            self.assertFalse(order.is_fulfilled)
            self.assertTrue(order.is_waiting)
