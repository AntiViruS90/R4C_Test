import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime
from django.db import DatabaseError
from .models import Robot

import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Count


# Create your views here.
@csrf_exempt
def create_robot(req):
    """
        Description:
        ------------
            Handle the creation of a new robot entry in the database.

            This view function processes POST requests to create a new robot
            with the specified model, version, and creation date. It expects
            a JSON payload with the required fields.

        Parameters:
        -----------
        req (HttpRequest): The HTTP request object containing the POST data.

        Returns:
        --------
        JsonResponse: A JSON response indicating the result of the operation.
                      - On success, returns a 201 status with a message and the robot ID.
                      - On failure, returns an appropriate error message and status code:
                        - 400 for missing fields, invalid JSON, or invalid date format.
                        - 405 if the request method is not POST.
                        - 500 for database errors or unexpected exceptions.
    """
    if req.method == "POST":
        try:
            data = json.loads(req.body)

            model = data.get('model')
            version = data.get('version')
            created_str = data.get('created')

            if not model or not version or not created_str:
                return JsonResponse(
                    {'error': "Missing required fields"},
                    status=400
                )

            created = parse_datetime(created_str)
            if not created:
                return JsonResponse(
                    {'error': "Invalid date format"},
                    status=400
                )

            serial = f"{model}{version}"

            robot = Robot.objects.create(
                serial=serial,
                model=model,
                version=version,
                created=created
            )

            return JsonResponse(
                {
                    'message': "Robot created succesfully",
                    "id": robot.id
                },
                status=201
            )

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)
        except DatabaseError as e:
            return JsonResponse(
                {"error": f"Database error: {str(e)}"},
                status=500
            )
        except Exception as e:
            return JsonResponse(
                {"error": f"Unexpected error: {str(e)}"},
                status=500
            )
    else:
        return JsonResponse(
            {'error': "Only POST method is allowed"},
            status=405
        )


def generate_robot_summary(req):
    """
        Description:
        ------------
            Generate an Excel summary of robots created in the last week.

            This function retrieves all robots created in the past week and generates
            an Excel file summarizing the count of each model and version. The Excel
            file is then returned as an HTTP response for download.

        Parameters:
        -----------
            req (HttpRequest): The HTTP request object.

        Returns:
        -------
            HttpResponse:
            An HTTP response with the Excel file attached, containing the summary of robots created in the last week.
    """
    one_week_ago = datetime.datetime.now() - datetime.timedelta(days=7)
    robots = Robot.objects.filter(created__gte=one_week_ago)

    wb = Workbook()
    wb.create_sheet(title="Placeholder")
    wb.remove(wb["Placeholder"])

    models = robots.values('model').distinct()

    for model in models:
        model_name = model['model']
        model_robots = robots.filter(model=model_name)

        version_summary = model_robots.values('version').annotate(count=Count('id'))
        ws = wb.create_sheet(title=model_name)

        ws.append(['Модель', 'Версия', 'Количество за неделю'])

        for entry in version_summary:
            ws.append([model_name, entry['version'], entry['count']])

        for col in range(1, 4):
            column = get_column_letter(col)
            max_length = 0

            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=robot_summary_last_week.xlsx'

    wb.save(response)

    return response
